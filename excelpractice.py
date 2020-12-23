from openpyxl import *

wb = load_workbook("Book1.xlsx")

ws = wb["Sheet2"]

ws['C3'] = "testing"

ws1 = wb.create_sheet("Mysheet")

wb.save("Book1.xlsx")